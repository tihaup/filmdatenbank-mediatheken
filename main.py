#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mediathek Film Tracker — Arte × ZDF × ARD × 3sat × Letterboxd × MUBI

Ablauf:
    1. Filme aus den öffentlich-rechtlichen Mediatheken laden (MediathekViewWeb-API)
    2. Beste Sprachversion pro Film auswählen
    3. Mit TMDb + Letterboxd anreichern (Bewertung, Regie, Beschreibung)
    4. Gegen die bestehende Filmdatenbank abgleichen → nur NEUE Filme behalten
    5. MUBI-Integration (Watchlist-Abgleich, eigene Ratings, Auto-Regisseure)
    6. Excel-Reports + CSV erzeugen, Datenbank aktualisieren

Ausgabe-Konvention:
    output/  → alle Excel-Reports (zum Ansehen)
    data/    → Zustandsdaten (CSV; für Logik + GitHub)

Start lokal:
    export TMDB_KEY="..."   (PowerShell:  $env:TMDB_KEY="...")
    python main.py
"""


# %%
import sys
import subprocess
subprocess.run([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"])

# %% ── Imports und Konfiguration ─────────────────────────────────────────────
import os
import re
import json
import time
from datetime import datetime
from collections import defaultdict

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Domänen-Konfiguration kommt zentral aus config.py (eine Quelle der Wahrheit).
from config import (
    SENDER,                # {sender: suchthema}
    LIEBLINGSREGISSEURE,   # Liste bevorzugter Regisseur:innen
    FARBE_SEHR_GUT,        # Hex-Farben für die Bewertungsstufen
    FARBE_GUT,
    FARBE_OK,
    MIN_DURATION_SEC,      # Mindestlaufzeit (schließt Kurzfilme aus)
)

# MUBI-Integration (optional — Pipeline läuft auch ohne)
try:
    import mubi
    MUBI_VERFUEGBAR = True
except ImportError:
    MUBI_VERFUEGBAR = False

# ── Operative Parameter (Laufzeit-Verhalten, nicht Domäne) ──────────────────
MAX_PRO_SENDER = 150       # Wie viele Filme pro Sender maximal geladen werden
MAX_FILMS      = 500       # Gesamt-Limit für die Anreicherung (None = alle).
                          # Für lokale Tests klein setzen, z.B. 5.
LETTERBOXD_DELAY = 0.8    # Pause zwischen Letterboxd-Anfragen (nicht < 0.5)

# ── Ordner und Pfade ────────────────────────────────────────────────────────
# Excel-Reports nach output/, Zustands-CSVs nach data/.
OUTPUT_DIR         = "output"
DATA_DIR           = "data"

DB_PFAD            = "data/filme_db.csv"        # kumulativ, wächst über die Zeit
LETZTE_WOCHE_PFAD  = "data/letzte_woche.csv"    # nur der jeweils letzte Lauf
AKTUELLE_WOCHE_CSV = "data/aktuelle_woche.csv"  # für direkte Ansicht auf GitHub
DB_EXCEL_PFAD      = "output/filme_datenbank.xlsx"   # Gesamt-DB als Excel

# TMDb API-Key (kostenlos von themoviedb.org/settings/api)
TMDB_KEY = os.environ.get("TMDB_KEY", "")
if not TMDB_KEY:
    raise ValueError("TMDB_KEY Umgebungsvariable nicht gesetzt!")

# Browser-ähnliche Session — reduziert Bot-Erkennung bei Letterboxd
session = requests.Session()
session.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
})


# %% ── Schritt 1: Filme aus den Mediatheken laden ───────────────────────────
def lade_mediathek_filme(sender, topic="film", size=MAX_PRO_SENDER,
                         duration_min=MIN_DURATION_SEC):
    """
    Lädt Filme eines Senders von MediathekViewWeb.

    Gibt eine Liste von Roh-Einträgen zurück. Ein Film kann hier mehrfach
    auftauchen (verschiedene Sprachfassungen) — das bereinigt Schritt 2.

    Wichtig: Content-Type MUSS text/plain sein, sonst antwortet die API mit 400.
    """
    query = {
        "queries": [
            {"fields": ["channel"], "query": sender},
            {"fields": ["topic"],   "query": topic},
        ],
        "sortBy":       "timestamp",   # neueste zuerst
        "sortOrder":    "desc",
        "future":       False,
        "offset":       0,
        "size":         size,
        "duration_min": duration_min,
    }

    antwort = requests.post(
        "https://mediathekviewweb.de/api/query",
        data=json.dumps(query),
        headers={"Content-Type": "text/plain"},  # MUSS text/plain sein
        timeout=30,
    )
    antwort.raise_for_status()

    data = antwort.json()
    if data.get("err"):
        raise ValueError(f"API-Fehler: {data['err']}")

    result_info = data.get("result", {})
    treffer     = result_info.get("results", [])
    return treffer


def lade_alle_sender():
    """
    Geht alle in SENDER konfigurierten Sender durch und sammelt die Rohdaten.
    Begrenzt pro Sender auf MAX_PRO_SENDER (Sicherheitsnetz zusätzlich zur API).
    """
    rohdaten = []
    for sender, topic in SENDER.items():
        print(f"Lade {sender}...")
        treffer = lade_mediathek_filme(sender, topic, size=MAX_PRO_SENDER)
        treffer = treffer[:MAX_PRO_SENDER]   # falls die API doch mehr liefert
        print(f"  → {len(treffer)} Treffer (begrenzt auf {MAX_PRO_SENDER})")
        rohdaten.extend(treffer)

    print(f"Gesamt: {len(rohdaten)} Einträge\n")
    return rohdaten


# %% ── Schritt 2: Beste Sprachversion pro Film auswählen ─────────────────────
def extrahiere_basis(titel):
    """
    Entfernt Versions-Suffixe und Mediathek-Zusätze am Titelende.
    'Minari (Originalversion mit Untertitel)' → 'Minari'
    'Match Point - Spielfilm, Großbritannien 2005' → 'Match Point'
    'Mo Hayder: Ritualmord (Französisch)' → 'Mo Hayder: Ritualmord'
    """
    basis = titel

    # 1. Sprach-/Versions-Suffixe in Klammern am Ende entfernen
    basis = re.sub(
        r"\s*\((Originalversion mit Untertitel|Originalversion|"
        r"mit Untertitel|Audiodeskription|Französisch|Englisch|"
        r"Italienisch|Spanisch|mit Gebärdensprache)\)\s*$",
        "",
        basis,
        flags=re.IGNORECASE,
    )

    # 2. Mediathek-Beschreibungszusatz nach " - Spielfilm" o.ä. abschneiden
    basis = re.sub(
        r"\s*[-–]\s*(Spielfilm|Fernsehfilm|Drama|Krimi|Komödie|Dokumentarfilm)\b.*$",
        "",
        basis,
        flags=re.IGNORECASE,
    )

    return basis.strip()


def versionstyp(titel):
    """
    Klassifiziert einen Titel nach Sprachfassung:
        'original' → (Originalversion mit Untertitel)  ← bevorzugt
        'deutsch'  → kein Suffix (deutsche Synchronfassung)  ← Fallback
        'skip'     → alle anderen Varianten  ← werden verworfen
    """
    if re.search(r"\(Originalversion mit Untertitel\)", titel, re.IGNORECASE):
        return "original"
    if re.search(r"\(mit Untertitel\)|\(Audiodeskription\)|\(Originalversion\)",
                 titel, re.IGNORECASE):
        return "skip"
    return "deutsch"


def waehle_beste_version(filme):
    """
    Wählt pro (Basis-Titel, Sender) genau eine Version:
        1. Originalversion mit Untertitel, sonst
        2. deutsche Fassung.
    Bei Mehrfachausstrahlung wird der neueste Eintrag (höchster timestamp) behalten.
    Derselbe Film bei verschiedenen Sendern bleibt als getrennte Einträge erhalten.
    """
    gruppen = defaultdict(dict)   # {(basis_titel, sender): {'original':…, 'deutsch':…}}

    for film in filme:
        basis  = extrahiere_basis(film["title"])
        sender = film.get("channel", "")
        typ    = versionstyp(film["title"])

        if typ == "skip":
            continue

        vorhandener = gruppen[(basis, sender)].get(typ)
        if vorhandener is None or film["timestamp"] > vorhandener["timestamp"]:
            gruppen[(basis, sender)][typ] = film

    ergebnis = []
    anzahl_original = 0
    anzahl_deutsch  = 0

    for versionen in gruppen.values():
        if "original" in versionen:
            ergebnis.append(versionen["original"])
            anzahl_original += 1
        elif "deutsch" in versionen:
            ergebnis.append(versionen["deutsch"])
            anzahl_deutsch += 1

    print(f"Rohdaten gesamt              : {len(filme)}")
    print(f"Eindeutige Basis-Titel+Sender: {len(gruppen)}")
    print(f"  davon Originalversion+UT   : {anzahl_original}")
    print(f"  davon deutsche Fassung     : {anzahl_deutsch}\n")
    return ergebnis


# %% ── Schritt 3: TMDb- und Letterboxd-Anreicherung ─────────────────────────
def hole_tmdb(deutscher_titel):
    """
    Sucht einen Film auf TMDb anhand des deutschen Titels.

    Rückgabe: (tmdb_id, original_titel, jahr, beschreibung)
    oder      (None, None, None, None) wenn nicht gefunden ODER bei API-Fehler.
    """
    try:
        antwort = requests.get(
            "https://api.themoviedb.org/3/search/movie",
            params={
                "api_key":  TMDB_KEY,
                "query":    deutscher_titel,
                "language": "de-DE",
            },
            timeout=10,
        )
        antwort.raise_for_status()
    except requests.RequestException as fehler:
        print(f"  ⚠ TMDb-Fehler bei '{deutscher_titel}': {fehler}")
        return None, None, None, None

    ergebnisse = antwort.json().get("results", [])
    if not ergebnisse:
        return None, None, None, None

    treffer = ergebnisse[0]
    tmdb_id        = treffer["id"]
    original_titel = treffer.get("original_title", deutscher_titel)
    jahr           = str(treffer.get("release_date", ""))[:4]
    beschreibung   = treffer.get("overview", "")
    return tmdb_id, original_titel, jahr, beschreibung


def hole_lb_filmseite_von_html(html, url):
    """
    Parst die Letterboxd-Filmseite. Die Bewertung etc. steckt im JSON-LD-Block
    (server-seitig gerendert), den wir hier auslesen.
    """
    soup = BeautifulSoup(html, "lxml")

    for script in soup.find_all("script", type="application/ld+json"):
        roh = script.string or ""
        roh = re.sub(r"/\*.*?\*/", "", roh, flags=re.DOTALL).strip()

        try:
            data = json.loads(roh)
        except (json.JSONDecodeError, TypeError):
            continue

        if data.get("@type") != "Movie":
            continue

        namen = []
        for regisseur in data.get("director", []):
            if regisseur.get("name"):
                namen.append(regisseur["name"])
        regie = ", ".join(namen)

        bewertung = data.get("aggregateRating", {})
        return {
            "lb_slug":     url.rstrip("/").split("/")[-1],
            "lb_title":    data.get("name", ""),
            "lb_year":     str(data.get("datePublished", ""))[:4],
            "lb_director": regie,
            "lb_desc":     data.get("description", ""),
            "lb_rating":   bewertung.get("ratingValue"),
            "lb_votes":    bewertung.get("ratingCount"),
            "lb_url":      url,
        }
    return None


def hole_lb_ueber_tmdb(tmdb_id):
    """
    letterboxd.com/tmdb/{id}/ leitet direkt zur passenden Filmseite weiter.
    So umgehen wir die JavaScript-Suche von Letterboxd.
    """
    url = f"https://letterboxd.com/tmdb/{tmdb_id}/"
    try:
        antwort = session.get(url, timeout=15, allow_redirects=True)
        antwort.raise_for_status()
    except requests.RequestException:
        return None
    return hole_lb_filmseite_von_html(antwort.text, antwort.url)


def hole_lb_daten(film):
    """
    Komplette Anreicherung für einen Film: Titel bereinigen → TMDb → Letterboxd.
    Fällt auf die TMDb-Beschreibung zurück, falls Letterboxd keine liefert.
    """
    titel = extrahiere_basis(film.get("title", ""))

    time.sleep(0.3)
    tmdb_id, _, _, tmdb_beschreibung = hole_tmdb(titel)
    if not tmdb_id:
        return {}

    time.sleep(LETTERBOXD_DELAY)
    lb_daten = hole_lb_ueber_tmdb(tmdb_id)

    if lb_daten and not lb_daten.get("lb_desc"):
        lb_daten["lb_desc"] = tmdb_beschreibung
    return lb_daten or {}


def reichere_alle_an(gefilterte_filme):
    """
    Läuft die komplette Anreicherungs-Pipeline für jeden Film durch.
    Gibt eine Liste von Ergebnis-Dicts zurück (Quelle + Letterboxd kombiniert).
    """
    kandidaten = sorted(gefilterte_filme,
                        key=lambda f: f.get("duration", 0),
                        reverse=True)

    if MAX_FILMS:
        kandidaten = kandidaten[:MAX_FILMS]

    print(f"Starte Anreicherung für {len(kandidaten)} Filme...")
    print(f"Geschätzte Dauer: {len(kandidaten) * 1.1 / 60:.1f} Minuten\n")

    ergebnisse = []
    for i, film in enumerate(kandidaten, 1):
        titel = extrahiere_basis(film.get("title", ""))
        dauer = f"{film.get('duration', 0) // 60} min"
        print(f"[{i:3d}/{len(kandidaten)}] {titel[:45]:45s} ({dauer})", end=" ")

        lb_daten = hole_lb_daten(film)

        if lb_daten:
            rating = lb_daten.get("lb_rating", "–")
            regie  = (lb_daten.get("lb_director") or "–")[:25]
            jahr   = lb_daten.get("lb_year", "–")
            print(f"→ ★ {rating}  {jahr}  {regie}")

            lb_daten["lb_title"] = titel  # Nutzt den bereinigten deutschen Titel

        else:
            print("→ nicht gefunden")

        # Quell-Daten (Mediathek) + Letterboxd-Daten zusammenführen
        eintrag = {
            "quell_titel":     film.get("title", ""),
            "dauer_min":       film.get("duration", 0) // 60,
            "hinzugefuegt_am": datetime.fromtimestamp(
                                   film.get("timestamp", 0)
                               ).strftime("%d.%m.%Y"),
            "thema":           film.get("topic", ""),
            "mediathek_url":   film.get("url_website", ""),
            "channel":         film.get("channel", ""),
            **lb_daten,
        }
        ergebnisse.append(eintrag)

    gefunden = sum(1 for e in ergebnisse if e.get("lb_rating"))
    print(f"\n✓ Fertig: {gefunden} von {len(ergebnisse)} Filmen auf Letterboxd gefunden")

    mit_bewertung = [e for e in ergebnisse if e.get("lb_rating")]
    print(f"✓ {len(mit_bewertung)} Filme mit Letterboxd-Bewertung behalten\n")
    return mit_bewertung


# %% ── Schritt 4: Datenbank-Abgleich (nur neue Filme) ───────────────────────
def lade_bekannte_filme(pfad=DB_PFAD):
    """
    Liest die bestehende, kumulative Filmdatenbank und gibt ein Set bekannter
    (titel, sender)-Kombinationen zurück. Vergleich case-insensitiv.
    Fehlt die Datei (erster Lauf), ist das Set leer → alle neu.
    """
    if not os.path.exists(pfad):
        return set()

    df_alt = pd.read_csv(pfad)
    bekannt = set()
    for _, zeile in df_alt.iterrows():
        titel  = str(zeile.get("titel", "")).lower().strip()
        sender = str(zeile.get("sender", "")).lower().strip()
        bekannt.add((titel, sender))
    return bekannt


def ist_neuer_film(eintrag, bekannte):
    """True, wenn (titel, sender) noch nicht in der Datenbank steht."""
    titel  = (eintrag.get("lb_title") or eintrag.get("quell_titel", "")).lower().strip()
    sender = eintrag.get("channel", "").lower().strip()
    return (titel, sender) not in bekannte


def finde_neue_filme(ergebnisse):
    """
    Filtert die bewerteten Filme gegen die bestehende Datenbank.
    Gibt nur die Filme zurück, die dort noch nicht stehen.
    """
    bekannte_filme = lade_bekannte_filme()

    neue_filme = []
    for eintrag in ergebnisse:
        if ist_neuer_film(eintrag, bekannte_filme):
            neue_filme.append(eintrag)

    print(f"✓ {len(neue_filme)} neue Filme (von {len(ergebnisse)} mit Bewertung)\n")
    return neue_filme


def aktualisiere_datenbank(ergebnisse):
    """
    Aktualisiert die persistente Datenbank NACH dem Abgleich:
      1. filme_db.csv     : neue Zeilen anhängen, Duplikate entfernen (kumulativ)
      2. letzte_woche.csv : mit aktuellem Lauf überschreiben

    Speichert jetzt ALLE relevanten Felder (inkl. mubi_rating + Beschreibung),
    damit die Gesamt-Datenbank-Excel vollständig ist.
    """
    os.makedirs(DATA_DIR, exist_ok=True)

    zeilen = []
    for e in ergebnisse:
        zeilen.append({
            "titel":           e.get("lb_title") or e.get("quell_titel", ""),
            "sender":          e.get("channel", ""),
            "lb_director":     e.get("lb_director", ""),
            "lb_year":         e.get("lb_year", ""),
            "lb_rating":       e.get("lb_rating"),
            "mubi_rating":     e.get("mubi_rating"),
            "lb_votes":        e.get("lb_votes"),
            "dauer_min":       e.get("dauer_min"),
            "hinzugefuegt_am": e.get("hinzugefuegt_am", ""),
            "lb_desc":         e.get("lb_desc", ""),
            "mediathek_url":   e.get("mediathek_url", ""),
            "lb_url":          e.get("lb_url", ""),
            "datum_gesehen":   "",
        })
    df_neu = pd.DataFrame(zeilen)

    if os.path.exists(DB_PFAD):
        df_alt  = pd.read_csv(DB_PFAD)
        df_alle = pd.concat([df_alt, df_neu], ignore_index=True)
        df_alle = df_alle.drop_duplicates(subset=["titel", "sender"], keep="first")
    else:
        df_alle = df_neu
    df_alle.to_csv(DB_PFAD, index=False)

    df_letzte_woche = df_neu[["titel", "sender"]]
    df_letzte_woche.to_csv(LETZTE_WOCHE_PFAD, index=False)

    print(f"DB gesamt   : {len(df_alle)} Einträge")
    print(f"Diese Woche : {len(df_neu)} Einträge\n")


# %% ── Schritt 5: Excel-Report bauen ────────────────────────────────────────
# Spalten: (interner Schlüssel, Anzeigename) — eine Quelle für Reihenfolge + Titel
SPALTEN = [
    ("lb_title",       "Titel"),
    ("channel",        "Sender"),
    ("lb_year",        "Jahr"),
    ("lb_director",    "Regie"),
    ("lb_rating",      "LB ★"),
    ("mubi_rating",    "MUBI ★"),
    ("lb_votes",       "Stimmen"),
    ("dauer_min",      "Dauer (min)"),
    ("hinzugefuegt_am", "Hinzugefügt am"),
    ("lb_desc",        "Beschreibung"),
    ("mediathek_url",  "Mediathek-Link"),
    ("lb_url",         "Letterboxd-Link"),
]

# Spaltenbreiten
BREITEN = {
    "lb_title":       28,
    "channel":         8,
    "lb_year":         6,
    "lb_director":    22,
    "lb_rating":       8,
    "mubi_rating":     8,
    "lb_votes":       12,
    "dauer_min":       8,
    "hinzugefuegt_am": 14,
    "lb_desc":        50,
    "mediathek_url":  14,
    "lb_url":         14,
}


def bewertungsfarbe(rating):
    """Ordnet einer Letterboxd-Bewertung die passende Hintergrundfarbe zu."""
    if rating is None:
        return None
    if rating > 4.0:
        return FARBE_SEHR_GUT
    if rating >= 3.66:
        return FARBE_GUT
    if rating >= 3.4:
        return FARBE_OK
    return None


def _schreibe_kopfzeile(ws, rahmen):
    """Hilfsfunktion: formatierte Kopfzeile in ein Tabellenblatt schreiben."""
    header_fill = PatternFill("solid", fgColor="1C1C2E")
    for col, (_, anzeigename) in enumerate(SPALTEN, 1):
        zelle = ws.cell(row=1, column=col, value=anzeigename)
        zelle.font      = Font(name="Arial", bold=True, color="F5C842", size=10)
        zelle.fill      = header_fill
        zelle.alignment = Alignment(horizontal="center", vertical="center")
        zelle.border    = rahmen
    ws.row_dimensions[1].height = 22


def _schreibe_datenzeilen(ws, eintraege, rahmen, desc_teiler, mit_farbe=True):
    """
    Hilfsfunktion: schreibt die Datenzeilen in ein Tabellenblatt.

    desc_teiler : steuert die Zeilenhöhen-Schätzung.
    mit_farbe   : Farbcodierung nach Letterboxd-Rating an/aus
                  (Wochen-Excel: an, Gesamt-Datenbank: aus).
    """
    for zeile_nr, eintrag in enumerate(eintraege, 2):
        rating = eintrag.get("lb_rating")
        farbe  = bewertungsfarbe(rating) if mit_farbe else None
        fill   = PatternFill("solid", fgColor=farbe) if farbe else None

        for col, (key, _) in enumerate(SPALTEN, 1):
            wert = eintrag.get(key, "")

            # Zahlen sauber formatieren
            if key == "lb_rating" and wert:
                wert = round(float(wert), 2)
            elif key == "mubi_rating" and wert:
                wert = int(wert)
            elif key == "lb_votes" and wert:
                wert = int(wert)

            zelle = ws.cell(row=zeile_nr, column=col, value=wert)
            zelle.font      = Font(name="Arial", size=9, bold=(key == "lb_rating"))
            zelle.alignment = Alignment(vertical="center",
                                        wrap_text=(key == "lb_desc"))
            zelle.border    = rahmen
            if fill:
                zelle.fill = fill
            if key == "lb_rating":
                zelle.number_format = "0.00"
            if key == "mubi_rating":
                zelle.number_format = "0"
            if key == "lb_votes":
                zelle.number_format = "#,##0"

        # Zeilenhöhe grob an die Beschreibungslänge anpassen
        zeichen = len(str(eintrag.get("lb_desc") or ""))
        zeilen  = max(1, -(-zeichen // desc_teiler))   # Aufrunden
        ws.row_dimensions[zeile_nr].height = zeilen * 13 + 4


def _schreibe_blatt(ws, eintraege, rahmen, desc_teiler, mit_farbe=True):
    """Schreibt Kopfzeile + Daten + Spaltenbreiten + Fixierung in EIN Blatt."""
    _schreibe_kopfzeile(ws, rahmen)
    _schreibe_datenzeilen(ws, eintraege, rahmen, desc_teiler, mit_farbe=mit_farbe)

    for col, (key, _) in enumerate(SPALTEN, 1):
        ws.column_dimensions[get_column_letter(col)].width = BREITEN.get(key, 12)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SPALTEN))}1"
    ws.freeze_panes    = "A2"


def baue_excel(neue_filme, favoriten, watchlist_treffer):
    """
    Baut die Wochen-Excel (nach output/) mit bis zu drei Blättern:
      1. Hauptblatt           : alle NEUEN Filme, nach Bewertung sortiert + Farbe
      2. Lieblingsregisseure  : Teilmenge der neuen Filme
      3. MUBI Watchlist       : aktuell laufende Watchlist-Filme (nur wenn welche)
    Gibt den Dateinamen zurück.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    thin   = Side(style="thin", color="CCCCCC")
    rahmen = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── Hauptblatt ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Mediatheken × Letterboxd"
    sortiert = sorted(neue_filme,
                      key=lambda e: e.get("lb_rating") or 0,
                      reverse=True)
    _schreibe_blatt(ws, sortiert, rahmen, desc_teiler=80, mit_farbe=True)

    # Legende unter die Tabelle
    legende_zeile = len(sortiert) + 3
    legende = [
        (FARBE_SEHR_GUT, "★ > 4.0    — Sehr gut"),
        (FARBE_GUT,      "★ 3.66–4.0 — Gut"),
        (FARBE_OK,       "★ 3.4–3.65 — Okay"),
    ]
    for farbe_hex, text in legende:
        zelle = ws.cell(row=legende_zeile, column=1, value=text)
        zelle.fill = PatternFill("solid", fgColor=farbe_hex)
        zelle.font = Font(name="Arial", size=9)
        legende_zeile += 1

    # ── Zweites Blatt: Lieblingsregisseure ──────────────────────────────────
    ws2 = wb.create_sheet(title="Lieblingsregisseure")
    favoriten_sortiert = sorted(
        favoriten,
        key=lambda e: (e.get("lb_director") or "", -(e.get("lb_rating") or 0)),
    )
    _schreibe_blatt(ws2, favoriten_sortiert, rahmen, desc_teiler=45, mit_farbe=True)
    print(f"✓ Blatt 'Lieblingsregisseure' mit {len(favoriten_sortiert)} Filmen erstellt")

    # ── Drittes Blatt: MUBI-Watchlist-Treffer (bewertungsunabhängig) ────────
    if watchlist_treffer:
        ws3 = wb.create_sheet(title="MUBI Watchlist")
        watchlist_sortiert = sorted(
            watchlist_treffer,
            key=lambda e: (e.get("mubi_rating") or 0, e.get("lb_rating") or 0),
            reverse=True,
        )
        _schreibe_blatt(ws3, watchlist_sortiert, rahmen, desc_teiler=80, mit_farbe=True)
        print(f"✓ Blatt 'MUBI Watchlist' mit {len(watchlist_sortiert)} Filmen erstellt")

    # ── Speichern ───────────────────────────────────────────────────────────
    dateiname = f"output/mediatheken_letterboxd_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(dateiname)

    anzahl_sehr_gut = sum(1 for e in sortiert if (e.get("lb_rating") or 0) > 4.0)
    anzahl_gut      = sum(1 for e in sortiert if 3.66 <= (e.get("lb_rating") or 0) <= 4.0)
    anzahl_okay     = sum(1 for e in sortiert if 3.4 <= (e.get("lb_rating") or 0) < 3.66)
    print(f"✓ Gespeichert: {dateiname}")
    print(f"  {len(sortiert)} neue Filme   |   {anzahl_sehr_gut} sehr gut   "
          f"|   {anzahl_gut} gut   |   {anzahl_okay} okay\n")
    return dateiname


def baue_datenbank_excel(pfad=DB_PFAD):
    """
    Exportiert die komplette filme_db.csv als übersichtliche Excel (nach output/),
    sortiert nach Regie (A→Z), innerhalb je Regie nach Bewertung absteigend.
    Filme ohne Regie landen am Ende. Mit Kurzbeschreibung, ohne Farbcodierung.
    """
    if not os.path.exists(pfad):
        print("⚠ Keine filme_db.csv vorhanden — DB-Excel übersprungen")
        return

    df = pd.read_csv(pfad)

    # CSV-Zeilen in Dicts mit SPALTEN-Schlüsseln umwandeln, damit wir dieselben
    # Schreibfunktionen wie für die Wochen-Excel nutzen können. NaN → None.
    def sauber(wert):
        return None if pd.isna(wert) else wert

    eintraege = []
    for _, zeile in df.iterrows():
        eintraege.append({
            "lb_title":        sauber(zeile.get("titel")),
            "channel":         sauber(zeile.get("sender")),
            "lb_year":         sauber(zeile.get("lb_year")),
            "lb_director":     sauber(zeile.get("lb_director")),
            "lb_rating":       sauber(zeile.get("lb_rating")),
            "mubi_rating":     sauber(zeile.get("mubi_rating")),
            "lb_votes":        sauber(zeile.get("lb_votes")),
            "dauer_min":       sauber(zeile.get("dauer_min")),
            "hinzugefuegt_am": sauber(zeile.get("hinzugefuegt_am")),
            "lb_desc":         sauber(zeile.get("lb_desc")),
            "mediathek_url":   sauber(zeile.get("mediathek_url")),
            "lb_url":          sauber(zeile.get("lb_url")),
        })

    # Sortierung: Regie A→Z (leere ans Ende), dann Bewertung absteigend
    def sortier_schluessel(e):
        regie      = e.get("lb_director") or ""
        regie_leer = 1 if regie.strip() == "" else 0
        rating     = e.get("lb_rating") or 0
        return (regie_leer, regie.lower(), -rating)

    eintraege_sortiert = sorted(eintraege, key=sortier_schluessel)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    thin   = Side(style="thin", color="CCCCCC")
    rahmen = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Filmdatenbank"
    _schreibe_blatt(ws, eintraege_sortiert, rahmen, desc_teiler=80, mit_farbe=False)

    wb.save(DB_EXCEL_PFAD)
    print(f"✓ Datenbank-Excel: {DB_EXCEL_PFAD} ({len(eintraege_sortiert)} Filme)\n")


# %% ── Schritt 6: CSV für direkte Ansicht auf GitHub ─────────────────────────
def speichere_aktuelle_woche_csv(neue_filme):
    """
    Schreibt die neuen Filme als feste CSV nach data/ (jede Woche überschrieben).
    GitHub rendert CSVs direkt als Tabelle im Browser → schnell ansehbar.
    Gleiche Spalten wie die Excel (inkl. Beschreibung), ohne Formatierung.
    """
    os.makedirs(DATA_DIR, exist_ok=True)

    sortiert = sorted(neue_filme,
                      key=lambda e: e.get("lb_rating") or 0,
                      reverse=True)

    zeilen = []
    for eintrag in sortiert:
        zeile = {}
        for key, anzeigename in SPALTEN:
            wert = eintrag.get(key, "")
            if key == "lb_rating" and wert:
                wert = round(float(wert), 2)
            elif key == "mubi_rating" and wert:
                wert = int(wert)
            elif key == "lb_votes" and wert:
                wert = int(wert)
            zeile[anzeigename] = wert
        zeilen.append(zeile)

    df_woche = pd.DataFrame(zeilen)
    # None/NaN als leere Zelle schreiben (sonst zeigt GitHub "NaN" / "")
    df_woche = df_woche.fillna("")
    df_woche.to_csv(AKTUELLE_WOCHE_CSV, index=False, encoding="utf-8-sig")
    print(f"✓ {AKTUELLE_WOCHE_CSV} geschrieben ({len(df_woche)} Filme)\n")


# %% ── Lieblingsregisseure-Filter ───────────────────────────────────────────
def ist_lieblingsregisseur(regie, namensliste):
    """True, wenn die Regie eine:n aus namensliste enthält."""
    if not regie:
        return False
    regie_lower = regie.lower()
    for name in namensliste:
        if name.lower() in regie_lower:
            return True
    return False


# %% ── main(): orchestriert den gesamten Ablauf ─────────────────────────────
def main():
    print("=== Mediathek Film Tracker ===\n")

    # Ausgabe-Ordner sicherstellen (kein Absturz, falls sie fehlen)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

    # 1. Rohdaten laden
    rohdaten = lade_alle_sender()

    # 2. Beste Version pro Film
    gefilterte_filme = waehle_beste_version(rohdaten)

    # 3. Mit TMDb + Letterboxd anreichern (nur Filme mit Bewertung bleiben)
    ergebnisse = reichere_alle_an(gefilterte_filme)

    # 4. Gegen bestehende DB abgleichen → nur NEUE Filme
    #    (MUSS vor aktualisiere_datenbank laufen, sonst wären alle "schon bekannt")
    neue_filme = finde_neue_filme(ergebnisse)

    # 5. MUBI-Integration: Watchlist-Abgleich, eigene Ratings, Auto-Regisseure.
    #    Verändert 'ergebnisse' in-place (hängt 'mubi_rating' an); da neue_filme
    #    dieselben Objekte referenziert, erscheint MUBI ★ auch dort.
    watchlist_treffer = []
    auto_regisseure   = []
    if MUBI_VERFUEGBAR:
        try:
            ergebnisse, watchlist_treffer, auto_regisseure = \
                mubi.reichere_mit_mubi_an(ergebnisse)
        except Exception as fehler:
            print(f"⚠ MUBI-Integration übersprungen: {fehler}")

    # 6. Lieblingsregisseure = handgepflegte Liste (config) + automatische (MUBI)
    alle_regisseure = list(LIEBLINGSREGISSEURE) + auto_regisseure
    favoriten = []
    for e in neue_filme:
        if ist_lieblingsregisseur(e.get("lb_director"), alle_regisseure):
            favoriten.append(e)
    print(f"✓ {len(favoriten)} neue Filme von Lieblingsregisseuren\n")

    # 7. Ausgaben erzeugen (Wochen-Excel + CSV) — beide nur mit neuen Filmen
    baue_excel(neue_filme, favoriten, watchlist_treffer)
    speichere_aktuelle_woche_csv(neue_filme)

    # 8. Datenbank aktualisieren (mit ALLEN bewerteten Filmen). Läuft vor der
    #    DB-Excel, damit diese den neuesten Stand abbildet.
    aktualisiere_datenbank(ergebnisse)

    # 9. Gesamt-Datenbank als Excel (nach Regie sortiert, mit Beschreibung)
    baue_datenbank_excel()

    print("=== Fertig ===")


if __name__ == "__main__":
    main()
